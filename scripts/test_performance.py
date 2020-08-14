#!/usr/bin/env python3.6

# import rosbag
import rospy
import rospkg
import os
import glob
import rosbag_pandas
import matplotlib.pyplot as plt
import statistics as stat
from openpyxl import Workbook
import pandas as pd
import numpy as np
import time
from mpl_toolkits import mplot3d
from sklearn.linear_model import LinearRegression
from sklearn.linear_model import Ridge
from sklearn.gaussian_process import GaussianProcessRegressor
from sklearn.neural_network import MLPRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.linear_model import SGDRegressor
from sklearn.svm import SVR
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import GridSearchCV, GroupKFold
import pickle

from functions_postprocess import *


# Paths
rospack = rospkg.RosPack()
path = rospack.get_path('model_training')
dir_bags = path + '/bags'
dir_figs = path + '/figures/'
dir_models = path + '/models/'
dir_results = path + '/results/'

# Experiment name and configs
exp = 'Experiment6'
# configs = ['dwa1','dwa2','teb1','teb2']
configs = ['dwa1','teb1']

d_topics = ['density1','narrowness1']

xtopics = ['density1','d_density1','narrowness1','d_narrowness1','performance1']
# ytopics = ['safety2']
ytopics = ['performance2_2','safety2']

models = ['SVR_rbf','RF']

# Resamplesize and smoothing (rolling)
samplesize = 10
rolling = 100

use_n_exp = 15

t = time.time()

# Import Bag files into pandas
os.chdir(dir_bags)
files = dict()
df = dict()

X = dict()
y = dict()
groups = dict()
mean_lags = dict()

for ytopic in ytopics:
    X[ytopic] = dict()
    y[ytopic] = dict()
    groups[ytopic] = dict()
    mean_lags[ytopic] = dict()
    for config in configs:
        df[config] = dict()

        files[config] = sorted(glob.glob("%s*%s.bag"%(exp,config)))
        lags = []
        n_exp = len(files[config])
        print('n_exp = %s, number of runs found'%n_exp)
        if 'use_n_exp' in locals():
            n_exp = use_n_exp
        print('n_exp = %s, number of runs used'%n_exp)

        for idx in range(n_exp):
            df[config][idx] = import_bag(files[config][idx],samplesize,rolling)
            df[config][idx] = add_derivs(df[config][idx],d_topics)

            df[config][idx] = df[config][idx].iloc[(int(4000/samplesize)):]

            corrs, lag = corrs_lags(df[config][idx],xtopics,ytopic,samplesize)
            lags.append(lag)
            print('idx = %s, corrs = %s, \tlags = %s'%(idx,corrs,lags[-1]))
            # df_shift = shift_lags(df[config][idx],xtopics,lags[config])
            # df_shift1 = shift_lags(df[config][idx],xtopics,-1*lags[config])
        mean_lags[ytopic][config] = np.array(lags).mean(axis=0).astype(int)
        print(np.array(lags).mean(axis=0).astype(int))

        for idx in range(n_exp):
            df_shift = shift_lags(df[config][idx],xtopics,mean_lags[ytopic][config])
            if idx == 0:
                # Shifted
                X[ytopic][config] = df_shift[xtopics].values
                y[ytopic][config] = df_shift[ytopic].values
                groups[ytopic][config] = np.full(len(df_shift[ytopic].values),idx)
            else:
                X[ytopic][config] = np.concatenate((X[ytopic][config], df_shift[xtopics].values))
                y[ytopic][config] = np.concatenate((y[ytopic][config], df_shift[ytopic].values))
                groups[ytopic][config] = np.concatenate((groups[ytopic][config], np.full(len(df_shift[ytopic].values),idx)))


# Plot example of signals
# for idx in range(n_exp):
# # for idx in [1]:
#     for config in configs:
#         topics = [[df[config][idx]['safety2'],df[config][idx]['performance2_2']],[df[config][idx]['narrowness1'],df[config][idx]['density1'],df[config][idx]['performance1']]]
#         titles = ['Experiment with config = %s'%(config),'Quality Attributes','Environment metrics']
#         xlabel = 'Time [s]'
#         ylabel = ['Quality Attributes','Environment Metrics']
#         fig = graph21(topics, titles, xlabel, ylabel)
# plt.show(block=False)
# plt.show()


# Print all the files with idx
print('idx   File')
for idx in range(n_exp):
    print('%-5s %-s'%(idx, files[configs[0]][idx]))


m = dict()
# models = ['SVR_rbf','SVR_poly']

# scoring_metric = 'neg_mean_squared_error'
scoring_metric = 'r2'

m['SVR_linear'] = GridSearchCV(SVR(kernel='linear',gamma=0.1),
                     param_grid={"C": [1, 25, 50, 75, 100, 125, 150],
                                 "gamma": np.logspace(-2, 2, 5)}, 
                     n_jobs=-1)
                     # scoring=scoring_metric)
m['SVR_rbf'] = GridSearchCV(SVR(kernel='rbf',gamma=0.1),
                     param_grid={"C": [1, 25, 50, 75, 100, 125, 150],
                                 "gamma": np.logspace(-2, 2, 5)}, 
                     n_jobs=-1)
                     # scoring=scoring_metric)
m['SVR_poly'] = GridSearchCV(SVR(kernel='poly',gamma=0.1),
                     param_grid={"C": [50, 100, 150, 200, 250],
                                 # "gamma": np.logspace(-2, 2, 5),
                                 "degree":[3,4]}, 
                     n_jobs=-1)
                     # scoring=scoring_metric)
# Random forrest
m['RF'] = GridSearchCV(RandomForestRegressor(),
                     param_grid={"n_estimators": [5, 10, 50, 100, 150, 200]}, 
                     n_jobs=-1)
                     # scoring=scoring_metric)
# Stochastic Gradient Descent
m['SGD'] = GridSearchCV(SGDRegressor(),
                     param_grid={"loss": ['squared_loss','huber','epsilon_insensitive'],
                                 "penalty":['l2','elasticnet'],
                                 "epsilon":[0.01,0.05,0.1,0.2]}, 
                     n_jobs=-1)
                     # scoring=scoring_metric)


# Train loop
os.chdir(dir_results)
for ytopic in ytopics:
    wb = Workbook()
    resultsfilename = 'Train_%s_%s.xlsx'%(ytopic,exp)
    for config in configs:
        print('Training for %s'%config)

        # Create sheet for config
        ws = wb.create_sheet(title=config)
        
        # Loop train models
        for modelname in models:
            # for idx in range(n_exp):
            gkf = GroupKFold(n_splits=n_exp)
            
            # fig, ax = plt.subplots():

            idm = 0
            best_score = -10
            for train, test in gkf.split(X[ytopic][config], y[ytopic][config], groups=groups[ytopic][config]):
                # Fit model
                m[modelname].fit(X[ytopic][config][train],y[ytopic][config][train])
                # Score 
                r2score = m[modelname].score(X[ytopic][config][test],y[ytopic][config][test])
                r2score_own = m[modelname].score(X[ytopic][config][train],y[ytopic][config][train])
                
                if r2score > best_score:
                    best_filename = "%s_%s_%s_%s_best.pkl"%(ytopic,modelname,config,idm)
                    best_filepath = dir_models + best_filename
                    model_best = m[modelname]
                    best_score = r2score
                    
                pkl_filename = dir_models + "%s_%s_%s_%s.pkl"%(ytopic,modelname,config,idm)
                with open(pkl_filename, 'wb') as file:
                    pickle.dump(m[modelname], file)

                print('Qa:%s, idm:%s, modelname:%s test score:%s, own score:%s'%(ytopic,idm,modelname,round(r2score,5),round(r2score_own,5)))
                
                # Parameters to results sheets
                # column names, only on top
                if idm == 0:
                    parnames = []
                    for x in m[modelname].best_params_:
                        parnames.append(x)
                    ws.append(['Model name','Trainset number','Test score','Own score'] + parnames)
                # Parameter values
                pars = []
                for x in m[modelname].best_params_:
                    pars.append(str(m[modelname].best_params_[x]))
                ws.append(["%s_%s_%s_%s.pkl"%(ytopic,modelname,config,idm), idm, r2score, r2score_own] + pars)
                wb.save(filename = resultsfilename)
                
                idm += 1

            print('Best model: %s'%best_filename)
            ws.append(['Best model'])
            ws.append([best_filename])
            ws.append(['Mean lags'])
            ws.append([str(mean_lags[ytopic][config])])
            with open(best_filepath, 'wb') as file:
                pickle.dump(model_best, file)

            # n=wb.sheetnames
            # wb.remove(wb["Sheet"])
            wb.save(filename = resultsfilename)

            

elapsed = time.time() - t
print('Elapsed time = %s'%elapsed)

